import connetDB
import pymysql
import openpyxl


def createTable():
    try:
        db = connetDB.connect_db()
        # 创建游标对象cursor
        cursor = db.cursor()
        # 使用execute()方法执行sql，如果表存在则删除,算了不删了
        cursor.execute('drop table if EXISTS skills')
        # 创建表的sql
        sql = '''
            create table skills(
            id int primary key auto_increment,
            name varchar(50) not null,
            damage int not null,
            classes int not null,
            hit int not null,
            PP int not null,
            attribute varchar(20),
            description varchar(500),
            priority int not null
             )
        '''

        cursor.execute(sql)

    except:
        print('创建表失败')
    finally:
        # 关闭数据库连接
        db.close()


# 已经创建再运行就会报错createTable()

# 插入，未完成版，没有参数化，参数化8个参数
# 名称name,伤害damage,类别classes,命中率hit,使用次数PP,属性attribute,描述description,优先级priority
def insert_skill(name="", damage=0, classes=0, hit=100, pp=0, attribute="", description="", priority=0):
    db = connetDB.connect_db()
    # 创建游标对象cursor
    cursor = db.cursor()
    sql = '''
        INSERT INTO skills(name,damage,classes,hit,PP,attribute,description,priority) 
        VALUES(%s,%s,%s,%s,%s,%s,%s,%s);'''
    # cursor.execute(sql, ('地震', 100, 0, 100, 10, '地面', '除自身以外场上全部可以攻击到的宝可梦', 0))
    cursor.execute(sql, (name, damage, classes, hit, pp, attribute, description, priority))
    db.commit()  # 这句一定要有，提交事务
    db.close()


# insert_skill("地震", 100, 0, 100, 10, "地面", "除自身以外场上全部可以攻击到的宝可梦", 0)
# insert_skill("电光束", 130, 1, 100, 10, "电", "第１回合收集电力提高特攻，第２回合将高压的电力发射出去。下雨天气时能立刻发射。", 0)

# 批量插入数据
# list格式 [("电光束", 130, 1, 100, 10, "电", "第１回合收集电力提高特攻，第２回合将高压的电力发射出去。下雨天气时能立刻发射。", 0)]
def insert_skills(skills_list):
    if skills_list == None:
        print("列表为空")
        return
    db = connetDB.connect_db()
    cursor = db.cursor()
    sql = '''
            INSERT INTO skills(name,damage,classes,hit,PP,attribute,description,priority) 
            VALUES(%s,%s,%s,%s,%s,%s,%s,%s);'''

    try:
        # 执行sql语句
        cursor.executemany(sql, skills_list)
        # 提交事务
        db.commit()
        print('插入多条数据成功')
    except Exception as e:
        print(e)
        # 如果出现异常，回滚
        db.rollback()
        print('插入多条数据失败')
    finally:
        # 关闭数据库连接
        db.close()


'''
list=[("月亮之力",95,1,100,15,"妖精","借用月亮的力量攻击对手。有时会降低对手的特攻。",0),
      ("魔法闪耀",80,4,100,10,"妖精","妖精系aoe",0)]

insert_skills(list)
'''


def update_skill_damage(name="", damage=0):
    if name == "":
        print("不知道名字改什么改")
        return

    db = connetDB.connect_db()
    cursor = db.cursor()  # 游标
    sql = 'update skills set damage=%s where name=%s'
    try:
        cursor.execute(sql, (damage, name))
        db.commit()
        print('修改成功')
    except:
        print('修改失败')
        db.rollback()
    finally:
        db.close()


# update_skill_damage("地震",100)


def delete_skill(name=""):
    if name == "":
        print("不知道名字删什么删")
        return

    db = connetDB.connect_db()
    cursor = db.cursor()  # 游标
    sql = 'delete from skills where name=%s'
    try:
        cursor.execute(sql, (name))
        db.commit()
        print('删除成功')
    except:
        print('删除失败')
        db.rollback()
    finally:
        db.close()


# delete_skill(name="地震")

def search_byname(name=""):
    if name == "":
        print("不知道名字删什么删")
        return

    db = connetDB.connect_db()
    cursor = db.cursor()  # 游标
    sql = 'select * from skills where name=%s'

    cursor.execute(sql, name)
    results = cursor.fetchall()  # 获取查询结果
    print(type(results))  # tuple(元组)类型，如果用一句话概括，那tuple类型就是“只读”的list
    print(results)
    for row in results:
        num = row[0]
        name = row[1]
        damage = row[2]
        classes = row[3]
        hit = row[4]
        pp = row[5]
        attribute = row[6]
        description = row[7]
        priority = row[8]

        print("序号", num, "名称", name, "伤害", damage, "类别", classes, "命中率", hit, "使用次数", pp, "属性", attribute, "\n",
              "描述", description, "优先度", priority)  # 输出


# search_byname("地震")

# 从数据库批量读取数据写入excel
def exl_write():
    db = connetDB.connect_db()
    try:
        with db.cursor() as cursor:  # 创建游标，在这里conn.cursor()==cursor
            cursor.execute(
                'select * from skills'
            )
            excel_workbook = openpyxl.Workbook()  # 创建一个excel工作簿
            excel_sheet = excel_workbook.create_sheet("招式")  # 在工作簿中建立一个工作学生成绩表，作为当前要写入内容的表
            excel_sheet = excel_workbook.active  # 使用默认的sheet表，不新建表,切换当前把数据写入此表
            # 注意数据量比较少的时候用fetchall()，一行一行的读取，这里数据量少就直接fetchall()了

            # openpyxl操作excel是，行和列的索引从1开始，所以要+1， col_id, col_name表示索引和索引内容
            # 列索引，读取那些列的数据
            for col_id, col_name in enumerate(
                    ['id', 'name', 'damage', 'classes', 'hit', 'pp', 'attribute', 'description', 'priority']):
                excel_sheet.cell(1, col_id + 1, col_name)  # 往单元格写入数据，第1列表格的表头''写入
            # 写入数据
            for row_id, row_emp in enumerate(cursor.fetchall()):  # 获得每一行的数据
                for col_id, col_value in enumerate(row_emp):  # 把每一行的每一列写入，加2是因为原先表头+1的基础上，再加上表头这一行，所以要+1+1=+2
                    excel_sheet.cell(row_id + 2, col_id + 1, col_value)
        excel_workbook.save("宝可梦信息.xlsx")
    except:  # 捕获异常
        print("写入失败")  # 如果出现异常，打印错误信息
    finally:
        db.close()  # 无论如何都要关闭连接，节省资源占用


# exl_write()

# 数据从excel批量导入
def exl_read():
    # 表处理
    # 在有主键的情况下不可重复导入
    db = connetDB.connect_db()
    excel_workbook = openpyxl.load_workbook("输入.xlsx")  # 创建一个excel工作簿
    excel_sheet = excel_workbook.active  # 使用默认的sheet表，不新建表,切换当前把数据写入此表
    excel_data = []  # 所有数据的列表
    # 从第二行第一列开始取得数据，不需要第一行表头的数据
    # 注意行列数在openpyxl库中是从0开始计数的，max_row是所有数据的行数20，我们要的是第2行到第21行
    for row_id in range(2, excel_sheet.max_row + 1):
        values = []  # 这一行数据的列表
        for col_id in range(1, excel_sheet.max_column + 1):
            values.append(excel_sheet.cell(row_id, col_id).value)
        print(values)
        excel_data.append(values)  # 实现的2层的列表嵌套
    # print(excel_data)#打印看看读取了什么数据

    try:
        with db.cursor() as cursor:
            # cursor.executemany表示批量插入数据，批处理
            cursor.executemany(
                'insert into skills'
                '(id,name,damage,classes,hit,pp,attribute,description,priority)'
                'values'
                '(%s,%s,%s,%s,%s,%s,%s,%s,%s)',
                excel_data
            )
        db.commit()
        print("导入成功！")
    except pymysql.MySQLError as err:  # 捕获异常
        print(err)  # 如果出现异常，打印错误信息
        print("导入失败！")
    finally:
        db.close()  # 无论如何都要关闭连接，节省资源占用


exl_read()
